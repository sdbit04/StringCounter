import os
import re
from openpyxl import Workbook
from tkinter import Tk, Entry, Button, Label, messagebox


def count_occurrences_of_string(search_string, input_file_path):
    _in_str = "value\s+\({}\)".format(search_string)
    with open("Search_MMS-UE-S1AP-DI.log", "a") as log_ob:
        print(_in_str, file=log_ob)
    _in_str_pat = re.compile(_in_str)
    add_enodebe_id = "ENODEB ID:"
    add_date_time = "TIME AND DATE:"
    add_time_stamp = "TIMESTAMP:"
    add_event_name = "MESSAGE TYPE:"
    with open(input_file_path, 'r') as input_file_ob:
        lines = input_file_ob.readlines()
        uplink_id_v_data = {}
        for ind, line in enumerate(lines):
            if re.search(_in_str_pat, line):
                uplink_id = line.strip()
                enodeb_id = None
                date_time = None
                time_stamp = None
                event_name = None
                for back_ind in range(ind, ind-100, -1):
                    back_line = lines[back_ind]
                    if re.search(add_enodebe_id, back_line):
                        enodeb_id = back_line.split(":")[1].strip()
                    elif re.search(add_date_time, back_line):
                        date_time = back_line.split(":")[1].strip()
                    elif re.search(add_time_stamp, back_line):
                        time_stamp = back_line.split(":")[1].strip()
                    elif re.search(add_event_name, back_line):
                        event_name = back_line.split(":")[1].strip()
                uplink_id_v_data[uplink_id]=[enodeb_id, date_time, time_stamp,event_name]

    return uplink_id_v_data


def count_occurrence_in_dir(search_string, input_dir_path):
    filename_v_uplink = {}
    for cur_dir, dirs, files in os.walk(input_dir_path):
        for input_file in files:
            with open("Search_MMS-UE-S1AP-DI.log", "a") as log_ob:
                print("Reading file = {}".format(input_file), file=log_ob)
            input_file_path = os.path.join(cur_dir, input_file)
            _uplink_id_v_data = count_occurrences_of_string(search_string, input_file_path)

            if _uplink_id_v_data is not None:
                filename_v_uplink[input_file] = _uplink_id_v_data
            else:
                with open("Search_MMS-UE-S1AP-DI.log", "a") as log_ob:
                    print("\t No match for {} found in this file".format(search_string), file=log_ob)
    else:
        return filename_v_uplink


def write_to_excel(file_v_uplink_and_related_data):
    ouput_file_path = "output.xlsx"
    WB = Workbook()
    sheet_ob = WB.create_sheet()
    headers = ["File_name", "UPLINK_ID", "ENODB_ID", "DATE-TIME", "TIMESTAMP", "EVENT_NAME"]
    for ind, head in enumerate(headers, 1):
        sheet_ob.cell(row=1, column=ind, value=head)
    start_row = 1
    for file_key, value in file_v_uplink_and_related_data.items():
        for search_key, add_vaules in value.items():
            start_row += 1
            sheet_ob.cell(row=start_row, column=1, value=file_key)
            sheet_ob.cell(row=start_row, column=2, value=search_key)
            for add_value_ind, add_value in enumerate(add_vaules, start=3):
                sheet_ob.cell(row=start_row, column=add_value_ind, value=add_value)

    WB.save(ouput_file_path)


def get_input():
    root = Tk()
    root.geometry("900x100")
    root.title("Search-MME-UE-#-ID")
    l_search_string = Label(root, text="Please provide Search: MME-UE-#-ID", padx=10)
    l_search_string.grid(row=0, column=0)
    e1 = Entry(root, width=100)
    e1.grid(row=0, column=1)
    l_input_dir = Label(root, text="Please provide input_directory", padx=10)
    l_input_dir.grid(row=2, column=0)
    e2 = Entry(root, width=100)
    e2.grid(row=2, column=1)

    # messagebox_1 =



    def my_click():
        global input_string
        input_string = None
        input_string = e1.get()
        global input_dir
        input_dir = None
        input_dir = e2.get()
        messagebox.showinfo("Program started!!", "Program Started check logs")
        root.destroy()

    my_button = Button(root, text="Submit", command=my_click)
    my_button.grid(row=5, column=1)
    root.mainloop()
    return input_string, input_dir


if __name__ == "__main__":
    # input_directory = r"D:\D_drive_BACKUP\MENTOR\Utility\Search_a_string\841104_OUT"
    # search_string = "MME-UE-S1AP-ID"
    # input_file = r"D:\D_drive_BACKUP\MENTOR\Utility\Search_a_string\840426_OUT\EmilGeoDump_200914T120001_200914T121459_840426.bin.gz_decoded_0.txt"
    # from argparse import ArgumentParser
    # parser = ArgumentParser(description="Please provide the Search-String and then search_directory_path")
    # parser.add_argument("search_string", help="Please provide string you are searching for")
    # parser.add_argument("input_dir", help="Please provide search_directory")
    # arguments = parser.parse_args()
    # search_string = arguments.search_string
    # input_directory = arguments.input_dir
    # print("Searching for " + search_string)

    search_string, input_directory = get_input()
    with open("Search_MMS-UE-S1AP-DI.log", "w") as log_ob:
        print("Starting to search MMS-UE-#-DI at {} directory".format(input_directory), file=log_ob)
    # search_string = "MME-UE-S1AP-ID"
    # input_directory = r"D:\D_drive_BACKUP\MENTOR\Utility\Search_a_string\New folder"
    filename_v_uplink_and_related_data = count_occurrence_in_dir(search_string, input_directory)
    # print("filename_v_uplink_and_related_data = {}".format(filename_v_uplink_and_related_data))
    write_to_excel(filename_v_uplink_and_related_data)


